from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook


# Caminho para o diretório do perfil do Chrome (onde seu login está)
chrome_user_data_dir = r"C:\Users\Win10\AppData\Local\Google\Chrome\User Data"# Substitua pelo caminho correto

# Configuração do Selenium para se conectar ao navegador já aberto
options = webdriver.ChromeOptions()
options.add_experimental_option("debuggerAddress", "localhost:9222")  # Conecta à porta de depuração
options.add_argument(f"user-data-dir={chrome_user_data_dir}")  # Usar o diretório de dados do usuário

# Usando o WebDriver Manager para baixar o ChromeDriver automaticamente
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# Abra o WhatsApp Web
driver.get("https://web.whatsapp.com/")
print("Conexão estabelecida com o Chrome no modo de depuração.")
sleep(30)  # Aumente esse tempo se necessário

# Carregar a planilha Excel
file_path = "dados.xlsx"  # Substitua pelo caminho do seu arquivo Excel
workbook = load_workbook(file_path)
sheet = workbook.active

# Loop pelos contatos na planilha
for row in sheet.iter_rows(min_row=7, values_only=True):  # Comece na linha 2 para ignorar cabeçalhos
    nome = row[0]  # Primeira coluna: Nome
    contato = str(row[1])  # Segunda coluna: Contato

    # Verifique se o contato é "(Não tem contato)"
    if contato == "(Não tem contato)":
        print(f"Ignorando {nome}, pois o contato está marcado como '(Não tem contato)'.")
        continue

    mensagem = f"Olá {nome}! Esta é uma mensagem automática enviada com Selenium."

    # Construir a URL do número no WhatsApp Web
    url = f"https://web.whatsapp.com/send?phone={contato}&text={mensagem}"

    # Abrir a conversa com o número
    driver.get(url)

    # Aguarde a página carregar
    sleep(15)

    try:
        # Clique no botão de enviar
        send_button = driver.find_element(By.XPATH, "//span[@data-icon='send']")
        send_button.click()
        print(f"Mensagem enviada para {nome} ({contato}).")
    except Exception as e:
        print(f"Não foi possível enviar mensagem para {nome} ({contato}). Erro: {e}")

    # Aguarde um tempo antes de enviar para o próximo contato
    sleep(10)

# Finalização
print("Mensagens enviadas para todos os contatos válidos da planilha.")
driver.quit()


#"C:\Program Files\Google\Chrome\Application\chrome.exe" --remote-debugging-port=9222 --user-data-dir="C:\ChromeDebug"