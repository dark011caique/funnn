import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import re
import subprocess
import os
from time import sleep
import PySimpleGUI as sg
from openpyxl import load_workbook
import tkinter as tk

def is_valid_contact(contact):
    """
    Verifica se o texto fornecido é um número de contato válido no formato brasileiro.
    """
    # Padrão para números de telefone brasileiros: (XX) XXXXX-XXXX ou (XX) XXXX-XXXX
    pattern = r"^\(?\d{2}\)?\s?\d{4,5}-\d{4}$"
    return bool(re.match(pattern, contact))

def create_interface():
    # Configurar o tema do PySimpleGUI
    sg.theme('Reddit')

    # Definir o layout da janela
    tela_busca = [
        [sg.Text('Estabelecimentos comerciais')],
        [sg.Input(key='comercio')],
        [sg.Text('Cidade')],
        [sg.Input(key='local')],
        [sg.Button('Buscar')]
    ]

    # Criar a janela
    janela = sg.Window('Buscar', layout=tela_busca)

    # Loop de eventos
    while True:
        event, values = janela.read()
        if event == sg.WIN_CLOSED or event == 'Buscar':
            break

    # Fechar a janela
    janela.close()

    # Pegar os valores de entrada
    comercio = values['comercio']
    local = values['local']

    # Concatenar os valores e buscar no Google Maps
    busca = f'{comercio} {local}'

    return busca

def create_mensagem_interface():
    # Configurar o tema do PySimpleGUI
    sg.theme('Reddit')

    # Definir o layout da janela
    tela_busca = [
        [sg.Text('Descrição', font=("Arial", 16))],
        [sg.Multiline(size=(50, 5), key='descricao')],
        [sg.Button('enviar')]
    ]

    # Criar a janela
    janela = sg.Window('enviar', layout=tela_busca)

    # Loop de eventos
    while True:
        event, values = janela.read()
        if event == sg.WIN_CLOSED or event == 'enviar':
            break

    # Fechar a janela
    janela.close()

    # Pegar os valores de entrada
    conteudo= values['descricao']
    

    # Concatenar os valores e buscar no Google Maps
    busca = f'{conteudo} '

    return busca

def extrair_dados_google_maps(driver, planilha_path, buscar):
    """
    Extrai dados de nome e contato do Google Maps e salva em uma planilha Excel.

    Args:
        driver (webdriver): O WebDriver do Selenium configurado.
        planilha_path (str): Caminho da planilha Excel para salvar os dados.
        buscar (str): Termo de busca no Google Maps.
    """

    planilha = openpyxl.load_workbook(planilha_path)
    dados = planilha['contatos']

    # Acessar o Google Maps
    driver.get("https://www.google.com.br/maps/preview")
    sleep(2)

    # Busca no Google Maps
    escreve = driver.find_element(By.XPATH, '//*[@id="searchboxinput"]')
    sleep(1)
    escreve.send_keys(buscar)

    botao_pesquisa = driver.find_element(By.XPATH, '//*[@id="searchbox-searchbutton"]/span')
    botao_pesquisa.click()
    sleep(1)

    # Scroll para carregar todos os resultados
    last_height = driver.execute_script("return document.body.scrollHeight")
    quantidade_anterior = 0

    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        sleep(2)
        resultados = driver.find_elements(By.XPATH, '//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]/div')
        quantidade_atual = len(resultados)

        if quantidade_atual > quantidade_anterior:
            ultimo_elemento = resultados[-1]
            driver.execute_script("arguments[0].scrollIntoView();", ultimo_elemento)
            sleep(1)
            quantidade_anterior = quantidade_atual
        else:
            break

    # Loop para clicar em cada elemento e extrair dados
    elements = driver.find_elements(By.XPATH, '//a[@class="hfpxzc"]')
    total_elements = len(elements)

    for i in range(1, total_elements + 1):
        try:
            xpath = f'(//a[@class="hfpxzc"])[{i}]'
            sleep(2)
            element = driver.find_element(By.XPATH, xpath)
            element.click()
            sleep(1)

            # Extrair nome e contato
            nome = driver.find_element(By.XPATH, '//h1[text()]').text
            contato = "Não tem contato"

            for j in range(2, 6):
                try:
                    contato_elemento = driver.find_element(By.XPATH, f'(//div[@class="Io6YTe fontBodyMedium kR99db fdkmkc "])[{j}]')
                    sleep(1)
                    texto_contato = contato_elemento.text
                    if is_valid_contact(texto_contato):
                        contato = texto_contato
                        break
                except:
                    continue

            print(f"Nome: {nome}, Contato: {contato}")
            sleep(2)
            dados.append([nome, contato])
            planilha.save(planilha_path)
            sleep(1)

            # Voltar para a lista
            driver.find_element(By.XPATH, '//*[@id="omnibox-singlebox"]/div/div[1]/button/span').click()
            sleep(3)
        except Exception as e:
            print(f"Erro ao processar elemento {i}: {e}")
            continue

def enviar_mensagens_whatsapp(file_path, chrome_user_data_dir, conteudo):
    """
    Envia mensagens para contatos armazenados em uma planilha Excel usando o WhatsApp Web.

    Args:
        file_path (str): Caminho da planilha Excel com os contatos.
        chrome_user_data_dir (str): Diretório de dados do usuário do Chrome.
        conteudo (str): Conteúdo da descrição inserida.
    """

    # Configuração do Selenium para se conectar ao navegador já aberto
    options = webdriver.ChromeOptions()
    options.add_experimental_option("debuggerAddress", "localhost:9222")  # Conecta à porta de depuração
    options.add_argument(f"user-data-dir={chrome_user_data_dir}")  # Usar o diretório de dados do usuário

    # Usando o WebDriver Manager para baixar o ChromeDriver automaticamente
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    # Abra o WhatsApp Web
    driver.get("https://web.whatsapp.com/")
    sleep(30)  # Aumente esse tempo se necessário

    # Carregar a planilha Excel
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Loop pelos contatos na planilha
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Comece na linha 3 para ignorar cabeçalhos
        nome = row[0]  # Primeira coluna: Nome
        contato = str(row[1])  # Segunda coluna: Contato

        # Verifique se o contato é "(Não tem contato)"
        if contato == "(Não tem contato)":
            print(f"Ignorando {nome}, pois o contato está marcado como '(Não tem contato)'.")
            continue

        # AQUI ESCREVE A MENSAGEM 
        mensagem = f"Olá {nome}! {conteudo}"

        # Construir a URL do número no WhatsApp Web
        url = f"https://web.whatsapp.com/send?phone={contato}&text={mensagem}"

        # Abrir a conversa com o número
        driver.get(url)

        # Aguarde a página carregar
        sleep(10)

        try:
            # Clique no botão de enviar
            send_button = driver.find_element(By.XPATH, "//span[@data-icon='send']")
            send_button.click()
            print(f"Mensagem enviada para {nome} ({contato}).")
        except Exception as e:
            print(f"Não foi possível enviar mensagem para {nome} ({contato}). Erro: {e}")

        # Aguarde um tempo antes de enviar para o próximo contato
        sleep(10)

    # Finalização
    print("Mensagens enviadas para todos os contatos válidos da planilha.")
    driver.quit()

def criar_planilha(buscar):
    """ Cria a planilha nomeada com base na busca e evita sobrescrita. """
    base_name = f"clientes_{buscar}" if buscar else "clientes"
    planilha_path = f"{base_name}.xlsx"
    contador = 1

    while os.path.exists(planilha_path):
        planilha_path = f"{base_name}_{contador}.xlsx"
        contador += 1

    planilha = openpyxl.Workbook()
    dados = planilha.active
    dados.title = "contatos"
    dados.append(["nomes", "contato"])
    planilha.save(planilha_path)

    return planilha_path  # Retorna o nome do arquivo criado

def main():
    """
    Função principal para executar o script de extração e envio de mensagens.
    """
    # Executar a interface primeiro
    buscar = create_interface()
    conteudo = create_mensagem_interface()

    planilha_path = criar_planilha(buscar)  # Cria a planilha com base na busca
    chrome_user_data_dir = r"C:\Users\Win10\AppData\Local\Google\Chrome\User Data"  # Substitua pelo caminho correto
    chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

    chrome_command = f'"{chrome_path}" --remote-debugging-port=9222 --user-data-dir="C:\\ChromeDebug"'

    try:
        print("Abrindo o navegador no modo de depuração...")
        subprocess.Popen(chrome_command, shell=True)
    except Exception as e:
        print(f"Erro ao abrir o Chrome: {e}")
        return
    
    sleep(2)

    # Configuração do Selenium para se conectar ao navegador já aberto
    options = webdriver.ChromeOptions()
    options.add_experimental_option("debuggerAddress", "localhost:9222")  # Conecta à porta de depuração
    options.add_argument(f"user-data-dir={chrome_user_data_dir}")  # Usar o diretório de dados do usuário

    # Usando o WebDriver Manager para baixar o ChromeDriver automaticamente
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    try:
        print("Extraindo dados do Google Maps...")
        extrair_dados_google_maps(driver, planilha_path, buscar)

        
        print("Finalizado pesquisa no Google Maps, vamos começar o envio de mensagem! ")
    

        print("Enviando mensagens no WhatsApp...")
        enviar_mensagens_whatsapp(planilha_path, chrome_user_data_dir, conteudo)

    finally:
        print("Processo concluído.")
        driver.quit()

if __name__ == "__main__":
    main()

