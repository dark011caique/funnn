import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import re
import subprocess
import os
from time import sleep
import customtkinter as ctk
from openpyxl import load_workbook

# Configurar tema do customtkinter
ctk.set_appearance_mode("dark")  # Modo escuro
ctk.set_default_color_theme("dark-blue")  # Tema azul escuro

def is_valid_contact(contact):
    """
    Verifica se o texto fornecido é um número de contato válido no formato brasileiro.
    """
    pattern = r"^\(?\d{2}\)?\s?\d{4,5}-\d{4}$"
    return bool(re.match(pattern, contact))

def listar_planilhas():
    """Lista todos os arquivos .xlsx no diretório atual."""
    return [f for f in os.listdir() if f.endswith('.xlsx')]

def atualizar_lista_planilhas(frame, planilhas):
    """Atualiza a lista de planilhas exibida na interface."""
    # Limpa o frame atual
    for widget in frame.winfo_children():
        widget.destroy()

    # Exibe as planilhas
    if not planilhas:
        ctk.CTkLabel(
            frame,
            text="Nenhuma planilha criada ainda.",
            font=("Arial", 14),
            text_color="white"
        ).pack(pady=10)
    else:
        for planilha in planilhas:
            ctk.CTkLabel(
                frame,
                text=planilha,
                font=("Arial", 14),
                text_color="white",
                fg_color="#3A3F40",
                corner_radius=5,
                width=300,
                height=40,
                anchor="w",
                padx=10
            ).pack(pady=5)

def extrair_dados_google_maps(driver, planilha_path, buscar):
    """
    Extrai dados de nome e contato do Google Maps e salva em uma planilha Excel.
    """
    planilha = openpyxl.load_workbook(planilha_path)
    dados = planilha['contatos']

    driver.get("https://www.google.com.br/maps/preview")
    sleep(2)

    escreve = driver.find_element(By.XPATH, '//*[@id="searchboxinput"]')
    sleep(1)
    escreve.send_keys(buscar)

    botao_pesquisa = driver.find_element(By.XPATH, '//*[@id="searchbox-searchbutton"]/span')
    botao_pesquisa.click()
    sleep(1)

    last_height = driver.execute_script("return document.body.scrollHeight")
    quantidade_anterior = 0

    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        sleep(2)
        resultados = driver.find_elements(By.XPATH, '//*[@id="QA0Szd"]/div/div/div[1]/div[2]/div/div[1]/div/div/div[1]/div[1]/div')
        quantidade_atual = len(resultados)

        if quantidade_atual > quantidade_anterior:
            ultimo_elemento = resultados[-1]
            driver.execute_script("arguments[0].scrollIntoView();", ultimo_elemento)
            sleep(1)
            quantidade_anterior = quantidade_atual
        else:
            break

    elements = driver.find_elements(By.XPATH, '//a[@class="hfpxzc"]')
    total_elements = len(elements)

    for i in range(1, total_elements + 1):
        try:
            xpath = f'(//a[@class="hfpxzc"])[{i}]'
            sleep(2)
            element = driver.find_element(By.XPATH, xpath)
            element.click()
            sleep(1)

            nome = driver.find_element(By.XPATH, '//h1[text()]').text
            contato = "Não tem contato"

            for j in range(2, 6):
                try:
                    contato_elemento = driver.find_element(By.XPATH, f'(//div[@class="Io6YTe fontBodyMedium kR99db fdkmkc "])[{j}]')
                    sleep(1)
                    texto_contato = contato_elemento.text
                    if is_valid_contact(texto_contato):
                        contato = texto_contato
                        break
                except:
                    continue

            print(f"Nome: {nome}, Contato: {contato}")
            sleep(2)
            dados.append([nome, contato])
            planilha.save(planilha_path)
            sleep(1)

            driver.find_element(By.XPATH, '//*[@id="omnibox-singlebox"]/div/div[1]/button/span').click()
            sleep(3)
        except Exception as e:
            print(f"Erro ao processar elemento {i}: {e}")
            continue

def enviar_mensagens_whatsapp(file_path, chrome_user_data_dir, conteudo):
    """
    Envia mensagens para contatos armazenados em uma planilha Excel usando o WhatsApp Web.
    """
    options = webdriver.ChromeOptions()
    options.add_experimental_option("debuggerAddress", "localhost:9222")
    options.add_argument(f"user-data-dir={chrome_user_data_dir}")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    driver.get("https://web.whatsapp.com/")
    sleep(30)

    workbook = load_workbook(file_path)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        nome = row[0]
        contato = str(row[1])

        if contato == "Não tem contato":
            print(f"Ignorando {nome}, pois o contato está marcado como 'Não tem contato'.")
            continue

        mensagem = f"Olá {nome}! {conteudo}"
        url = f"https://web.whatsapp.com/send?phone={contato}&text={mensagem}"

        driver.get(url)
        sleep(10)

        try:
            send_button = driver.find_element(By.XPATH, "//span[@data-icon='send']")
            send_button.click()
            print(f"Mensagem enviada para {nome} ({contato}).")
        except Exception as e:
            print(f"Não foi possível enviar mensagem para {nome} ({contato}). Erro: {e}")

        sleep(10)

    print("Mensagens enviadas para todos os contatos válidos da planilha.")
    driver.quit()

def criar_planilha(buscar):
    """Cria a planilha nomeada com base na busca e evita sobrescrita."""
    base_name = f"clientes_{buscar}" if buscar else "clientes"
    planilha_path = f"{base_name}.xlsx"
    contador = 1

    while os.path.exists(planilha_path):
        planilha_path = f"{base_name}_{contador}.xlsx"
        contador += 1

    planilha = openpyxl.Workbook()
    dados = planilha.active
    dados.title = "contatos"
    dados.append(["nomes", "contato"])
    planilha.save(planilha_path)

    return planilha_path

def main():
    """
    Função principal para executar o script de extração e envio de mensagens.
    """
    root = ctk.CTk()
    root.title("Sistema de Busca e Envio")
    root.geometry("800x500")
    root.configure(fg_color="#1C2526")

    # Frame lateral esquerdo para os campos de entrada
    left_frame = ctk.CTkFrame(root, width=400, fg_color="#2A2F30", corner_radius=0)
    left_frame.pack(side="left", fill="y", padx=10, pady=10)

    # Título
    ctk.CTkLabel(
        left_frame,
        text="Pesquisa e Descrição",
        font=("Arial", 20, "bold"),
        text_color="white"
    ).pack(pady=20)

    # Label e entrada para "Estabelecimentos comerciais"
    ctk.CTkLabel(
        left_frame,
        text="Estabelecimentos comerciais",
        font=("Arial", 14),
        text_color="white"
    ).pack(pady=(10, 0))
    comercio_entry = ctk.CTkEntry(
        left_frame,
        width=300,
        height=40,
        font=("Arial", 14),
        fg_color="#3A3F40",
        text_color="white",
        placeholder_text="Digite o tipo de comércio",
        placeholder_text_color="gray"
    )
    comercio_entry.pack(pady=10)

    # Label e entrada para "Cidade"
    ctk.CTkLabel(
        left_frame,
        text="Cidade",
        font=("Arial", 14),
        text_color="white"
    ).pack(pady=(10, 0))
    local_entry = ctk.CTkEntry(
        left_frame,
        width=300,
        height=40,
        font=("Arial", 14),
        fg_color="#3A3F40",
        text_color="white",
        placeholder_text="Digite a cidade",
        placeholder_text_color="gray"
    )
    local_entry.pack(pady=10)

    # Label e campo de texto para "Descrição"
    ctk.CTkLabel(
        left_frame,
        text="Descrição",
        font=("Arial", 14),
        text_color="white"
    ).pack(pady=(10, 0))
    descricao_text = ctk.CTkTextbox(
        left_frame,
        width=300,
        height=100,
        font=("Arial", 14),
        fg_color="#3A3F40",
        text_color="white"
    )
    descricao_text.pack(pady=10)

    # Frame direito para exibir as planilhas
    right_frame = ctk.CTkFrame(root, fg_color="#1C2526")
    right_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)

    # Título do frame de planilhas
    ctk.CTkLabel(
        right_frame,
        text="Planilhas Criadas",
        font=("Arial", 20, "bold"),
        text_color="white"
    ).pack(pady=20)

    # Frame para a lista de planilhas (com scroll)
    planilhas_frame = ctk.CTkScrollableFrame(right_frame, fg_color="#2A2F30", corner_radius=10)
    planilhas_frame.pack(fill="both", expand=True, padx=20)

    # Lista inicial de planilhas
    planilhas = listar_planilhas()
    atualizar_lista_planilhas(planilhas_frame, planilhas)

    # Botão "Executar"
    def executar():
        comercio = comercio_entry.get()
        local = local_entry.get()
        buscar = f"{comercio} {local}"
        conteudo = descricao_text.get("1.0", "end-1c").strip()

        if not buscar.strip() or not conteudo:
            ctk.CTkLabel(
                left_frame,
                text="Por favor, preencha todos os campos!",
                font=("Arial", 12),
                text_color="red"
            ).pack(pady=5)
            return

        planilha_path = criar_planilha(buscar)
        chrome_user_data_dir = r"C:\Users\Win10\AppData\Local\Google\Chrome\User Data"
        chrome_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"

        chrome_command = f'"{chrome_path}" --remote-debugging-port=9222 --user-data-dir="C:\\ChromeDebug"'

        try:
            print("Abrindo o navegador no modo de depuração...")
            subprocess.Popen(chrome_command, shell=True)
        except Exception as e:
            print(f"Erro ao abrir o Chrome: {e}")
            return

        sleep(2)

        options = webdriver.ChromeOptions()
        options.add_experimental_option("debuggerAddress", "localhost:9222")
        options.add_argument(f"user-data-dir={chrome_user_data_dir}")

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

        try:
            print("Extraindo dados do Google Maps...")
            extrair_dados_google_maps(driver, planilha_path, buscar)

            print("Finalizado pesquisa no Google Maps, vamos começar o envio de mensagem!")
            print("Enviando mensagens no WhatsApp...")
            enviar_mensagens_whatsapp(planilha_path, chrome_user_data_dir, conteudo)

        finally:
            print("Processo concluído.")
            driver.quit()

        # Atualiza a lista de planilhas após a criação
        planilhas = listar_planilhas()
        atualizar_lista_planilhas(planilhas_frame, planilhas)

    ctk.CTkButton(
        left_frame,
        text="Executar",
        command=executar,
        fg_color="#28A745",  # Verde
        hover_color="#218838",
        font=("Arial", 14, "bold"),
        text_color="white",
        width=300,
        height=40
    ).pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    main()
